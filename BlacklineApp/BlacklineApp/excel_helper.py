import openpyxl
import log_helper
import logging
from openpyxl import load_workbook
import time
#from ctypes import cdll
import os

#_sopen = cdll.msvcrt._sopen
#_close = cdll.msvcrt._close
#_SH_DENYRW = 0x10

def is_not_open(filename):
    wb = openpyxl.load_workbook(filename)
    try:
        wb.save(filename)
        return True
        #if not os.access(filename, os.F_OK):
        #    return True # file doesn't exist
        #h = _sopen(filename, 0, _SH_DENYRW, 0)
        #if h == 3:
        #    _close(h)
        #    logging.info('is_open open:' +filename)
        #    return True # file is not opened by anyone else
        #return False # file is already open
    except Exception as ex:
        logging.exception('is_open errr:' +str(ex))
        return False
    finally:
        wb.close()

def copy_sheet(sheetname,source_file_path,target_file_path):
     '''copy sheet '''
     wait_time = 300
     while wait_time>0:
         #source_file_opened = is_open(source_file_path)
         target_file_opened = is_not_open(target_file_path)
         if target_file_opened:
             break
         wait_time = wait_time - 2
         time.sleep(2)

     source_wb = load_workbook(source_file_path)
     target_wb = load_workbook(target_file_path)
     try: 
         sheet_names = target_wb.get_sheet_names()

         have_sheet = False
         for name in sheet_names:
            if sheetname == name:
                have_sheet = True
                break
         if not have_sheet:
                target_wb.create_sheet(title=sheet_name,index = index)
         #target_wb.save(file_path)
         #target_wb.close()

         source_ws = source_wb[sheetname]
         target_ws = target_wb[sheetname]
 
         #两个for循环遍历整个excel的单元格内容
         for i,row in enumerate(source_ws.iter_rows()):
          for j,cell in enumerate(row):
           target_ws.cell(row=i+1, column=j+1, value=cell.value)
     except Exception as ex:
         logging.info('copy_sheet :' +str(ex))
     finally:
         target_wb.save(target_file_path)
         target_wb.close()
         source_wb.close()



def write_list_data_to_excel(list_data,file_path,sheet_name):
    ''' write list data to excel '''
    wait_time = 300
    while wait_time>0:
        #source_file_opened = is_open(source_file_path)
        target_file_opened = is_not_open(file_path)
        if target_file_opened:
            break
        wait_time = wait_time - 2
        time.sleep(2)

    wb = openpyxl.load_workbook(file_path)
    try:
        sheet_report = wb.get_sheet_by_name(sheet_name) 
        if len(list_data) > 0:
            for row in list_data:
                sheet_report.append(row)
        return True
    except Exception as ex:
        return False
        logging.exception('set_colunm_name :' + str(ex))
    finally:
        wb.save(file_path)
        wb.close()

def create_excel_file(file_path):
    '''create excel file '''
    try:
        wb = openpyxl.Workbook()
        wb.save(file_path)
        return True
    except Exception as ex:
        logging.exception('create_excel_file :' + str(ex))
        return False

def create_sheet_name(file_path,sheet_name,index = None):
    wait_time = 300
    while wait_time>0:
        #source_file_opened = is_open(source_file_path)
        target_file_opened = is_not_open(file_path)
        if target_file_opened:
            break
        wait_time = wait_time - 2
        time.sleep(2)

    wb = openpyxl.load_workbook(file_path)
    try:
        sheet_names = wb.get_sheet_names()

        have_sheet = False
        for name in sheet_names:
            if sheet_name == name:
                have_sheet = True
                break
        if not have_sheet:
             wb.create_sheet(title=sheet_name,index = index)
             wb.save(file_path)
        return True
    except Exception as ex:
        logging.exception('create_sheet_name :' + str(ex))
        return False
    finally:
        wb.close()

        
def set_colunm_name(table_title,file_path,sheet_name):
    wait_time = 300
    while wait_time>0:
        #source_file_opened = is_open(source_file_path)
        target_file_opened = is_not_open(file_path)
        if target_file_opened:
            break
        wait_time = wait_time - 2
        time.sleep(2)

    wb = openpyxl.load_workbook(file_path)
    try:
        #tableTitle = ['userName', 'Phone', 'age', 'Remark']
        ws = wb.get_sheet_by_name(sheet_name) 
        for col in range(len(table_title)):
            c = col + 1
            ws.cell(row=1, column=c).value = table_title[col]
        return True
    except Exception as ex:
        return False
        logging.exception('set_colunm_name :' + str(ex))
    finally:
        wb.save(file_path)
        wb.close()

def write_row_data(data_list,file_path,sheet_name):
    wait_time = 300
    while wait_time>0:
        #source_file_opened = is_open(source_file_path)
        target_file_opened = is_not_open(file_path)
        if target_file_opened:
            break
        wait_time = wait_time - 2
        time.sleep(2)

    wb = openpyxl.load_workbook(file_path)
    try:
        sheet_report = wb.get_sheet_by_name(sheet_name) 
       
        sheet_report.append(data_list)
        return True
    except Exception as ex:
        return False
        logging.exception('set_colunm_name :' + str(ex))
    finally:
        wb.save(file_path)
        wb.close()

#def copy_sheet(sheetname,source_file_path,target_file_path):
# '''copy sheet '''
# source_wb = load_workbook(source_file_path)
# target_wb = load_workbook(target_file_path)
# try: 
#     sheet_names = target_wb.get_sheet_names()

#     have_sheet = False
#     for name in sheet_names:
#        if sheetname == name:
#            have_sheet = True
#            break
#     if not have_sheet:
#            target_wb.create_sheet(title=sheet_name,index = index)
#     #target_wb.save(file_path)
#     #target_wb.close()

#     source_ws = source_wb[sheetname]
#     target_ws = target_wb[sheetname]
 
#     #两个for循环遍历整个excel的单元格内容
#     for i,row in enumerate(source_ws.iter_rows()):
#      for j,cell in enumerate(row):
#       target_ws.cell(row=i+1, column=j+1, value=cell.value)
# except Exception as ex:
#     logging.info('copy_sheet :' +str(ex))
# finally:
#     target_wb.save(target_file_path)
#     target_wb.close()
#     source_wb.close()



#def write_list_data_to_excel(list_data,file_path,sheet_name):
#    ''' write list data to excel '''
#    wb = openpyxl.load_workbook(file_path)
#    try:
#        sheet_report = wb.get_sheet_by_name(sheet_name) 
#        if len(list_data) > 0:
#            for row in list_data:
#                sheet_report.append(row)
#        return True
#    except Exception as ex:
#        return False
#        log_helper.add_log('set_colunm_name :' + str(ex))
#    finally:
#        wb.save(file_path)
#        wb.close()

#def create_excel_file(file_path):
#    '''create excel file '''
#    try:
#        wb = openpyxl.Workbook()
#        wb.save(file_path)
#        return True
#    except Exception as ex:
#        log_helper.add_log('create_excel_file :' + str(ex))
#        return False

#def create_sheet_name(file_path,sheet_name,index = None):
#    wb = openpyxl.load_workbook(file_path)
#    try:
#        sheet_names = wb.get_sheet_names()

#        have_sheet = False
#        for name in sheet_names:
#            if sheet_name == name:
#                have_sheet = True
#                break
#        if not have_sheet:
#             wb.create_sheet(title=sheet_name,index = index)
#             wb.save(file_path)
#        return True
#    except Exception as ex:
#        log_helper.add_log('create_sheet_name :' + str(ex))
#        return False
#    finally:
#        wb.close()

        
#def set_colunm_name(table_title,file_path,sheet_name):
#    wb = openpyxl.load_workbook(file_path)
#    try:
#        #tableTitle = ['userName', 'Phone', 'age', 'Remark']
#        ws = wb.get_sheet_by_name(sheet_name) 
#        for col in range(len(table_title)):
#            c = col + 1
#            ws.cell(row=1, column=c).value = table_title[col]
#        return True
#    except Exception as ex:
#        return False
#        log_helper.add_log('set_colunm_name :' + str(ex))
#    finally:
#        wb.save(file_path)
#        wb.close()

#def write_row_data(data_list,file_path,sheet_name):
#    wb = openpyxl.load_workbook(file_path)
#    try:
#        sheet_report = wb.get_sheet_by_name(sheet_name) 
       
#        sheet_report.append(data_list)
#        return True
#    except Exception as ex:
#        return False
#        log_helper.add_log('set_colunm_name :' + str(ex))
#    finally:
#        wb.save(file_path)
#        wb.close()

